import openpyxl
import pickle


class Entry:
    num_right_guess = 0
    num_close_guess = 0
    num_wrong_guess = 0

    def __init__(self, word_in, answer_in):
        self.word = word_in
        self.answer = answer_in
        self.rank = 0
        # self.num_right_guess = 0
        # self.num_close_guess = 0
        # self.num_wrong_guess = 0

    def update_rank(self):
        self.rank = self.num_right_guess*3 + self.num_close_guess*2 + self.num_wrong_guess

def get_default():
    main_sheet = openpyxl.load_workbook('../../wokabelheft_zu_importieren.xlsx').active
    dict_entries = {}
    for i in range(1, main_sheet.max_row + 1):
        dict_entries[main_sheet['A{}'.format(i)].value] = Entry(main_sheet['A{}'.format(i)].value,
                                                                main_sheet['B{}'.format(i)].value)
    return dict_entries


def get_dict_from_excl_at(path):
    main_sheet = openpyxl.load_workbook(path).active
    dict_entries = {}
    for i in range(1, main_sheet.max_row + 1):
        dict_entries[main_sheet['A{}'.format(i)].value] = Entry(main_sheet['A{}'.format(i)].value,
                                                                main_sheet['B{}'.format(i)].value)
    return dict_entries


def add_data():
    # verify doubled entries
    pass


def get_non_seen(dictionary):
    non_seen_entries = {}
    for k in list(dictionary.keys()):
        card_rank = dictionary[k].num_right_guess*3 + dictionary[k].num_close_guess*2 + dictionary[k].num_wrong_guess
        if card_rank == 0:
            non_seen_entries[k] = dictionary[k]
    return non_seen_entries


def get_all_learning_order(dictionary):
    pass


def save_learning_state(dict_entries):  # add list_right_guesses, list_close_guesses, list_wrong_guesses
    f = open("dictionary.pkl", "wb")
    pickle.dump(dict_entries, f)
    f.close()


def save_learning_state_at(dict_entries, path):  # add list_right_guesses, list_close_guesses, list_wrong_guesses
    f = open(path, "wb")
    pickle.dump(dict_entries, f)
    f.close()

def load_learning_state(dict_path):
    with open(dict_path, 'rb') as handle:
        data = handle.read()
    return pickle.loads(data)

